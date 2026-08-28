from io import BytesIO
import json
import logging
import re
from time import monotonic
from zipfile import ZIP_DEFLATED, ZipFile
from fastapi import APIRouter, BackgroundTasks, Cookie, Depends, File, HTTPException, Query, Response, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import case, delete, func, insert, literal, or_, select, update
from sqlalchemy.orm import Session, selectinload
import xlsxwriter
from openpyxl import load_workbook
from app.db.session import get_db
from app.models.entities import AuditLog, Client, ClientChange, ClientStatus, Email, EmailReportExclusion, FtpImportEvent, Import, ImportIssue, Phone, TradePlace
from app.schemas.client import BulkUpdate, ClientDetail, ClientListItem, PagedClients
from app.services.import_tasks import create_import_task, get_import_task, run_import_task
from app.services.ftp_import import get_ftp_settings, get_ftp_status, mark_ftp_pending, run_ftp_import, save_ftp_settings, test_connection
from app.services.ftp_scheduler import refresh_ftp_schedule
from app.services.settings_auth import COOKIE_NAME, authenticate_settings, is_settings_authenticated, logout_settings
from app.services.client_changes import payload_signature, record_client_change
from app.services.normalization import extract_emails

router = APIRouter(prefix="/api", tags=["clients"])
import_logger = logging.getLogger("clients.import")

EXPORT_COLUMNS = {
    "name": ("Наименование", lambda client: client.name),
    "company": ("Фирма", lambda client: client.company or ""),
    "manager": ("Менеджер", lambda client: client.manager or ""),
    "phones": ("Телефоны", lambda client: "\n".join(sorted({item.phone for item in client.phones}))),
    "emails": ("Email", lambda client: "\n".join(sorted({item.email.strip() for item in client.emails if item.email.strip()}))),
}

DEFAULT_EMAIL_REPORTS = [
    {"name": "Корпоративные клиенты", "price_types": ["Корпоративные"], "buyer_types": ["Корпоративный"], "managers": []},
    {"name": "Розничные клиенты", "price_types": ["Розничные"], "buyer_types": ["Розница"], "managers": []},
    {"name": "Пашута ОПТ", "price_types": ["Оптовые"], "buyer_types": ["Оптовик"], "managers": ["Пашута М.С.", "Пашута М.С. (Ростов)"]},
    {"name": "Родина, Самойлова", "price_types": ["Оптовые"], "buyer_types": ["Оптовик"], "managers": ["Родина", "Самойлова", "Родина Е.В. (Ростов)"]},
    {"name": "Трошина, Гончарова", "price_types": ["Оптовые"], "buyer_types": ["Оптовик"], "managers": ["Трошина Лариса"]},
    {"name": "Шакулова", "price_types": ["Оптовые"], "buyer_types": ["Оптовик"], "managers": ["Шакулова Екатерина"]},
    {"name": "Суркова, Ромащенко, Бабушкина, Новожилова", "price_types": ["Оптовые"], "buyer_types": ["Оптовик"], "managers": ["Суркова Н.", "Ромащенко Екатерина", "Бабушкина Виктория", "Новожилова М."]},
    {"name": "Селянкина, Королева", "price_types": ["Оптовые"], "buyer_types": ["Оптовик"], "managers": ["Селянкина Татьяна", "Королева Светлана"]},
]

RETAIL_EMAIL_REPORT_NAME = "Розничные клиенты"
RETAIL_EMAIL_REPORT_EXCLUDED_ABBREVIATIONS = (
    "ООО", "ИП", "ГУП", "МОУ", "ЗАО", "ВАО", "АО", "МДОБУ",
    "НПО", "ТД", "ОМОН", "МУП", "ОАО", "СК", "ТНП",
)
RETAIL_EMAIL_REPORT_EXCLUDED_NAME_PATTERN = re.compile(
    rf"\b(?:{'|'.join(RETAIL_EMAIL_REPORT_EXCLUDED_ABBREVIATIONS)})\b",
    re.IGNORECASE,
)
EMAIL_EXCLUSION_CATEGORIES = {"unsubscribed", "problematic"}


def require_settings_auth(clients_settings_session: str | None = Cookie(None)) -> None:
    if not is_settings_authenticated(clients_settings_session):
        raise HTTPException(status_code=401, detail="Введите пароль для доступа к настройкам")


@router.post("/settings/auth")
def settings_login(payload: dict, response: Response):
    token = authenticate_settings(str(payload.get("password", "")))
    if token is None:
        raise HTTPException(status_code=401, detail="Неверный пароль")
    response.set_cookie(COOKIE_NAME, token, httponly=True, secure=True, samesite="strict", max_age=12 * 60 * 60)
    return {"authenticated": True}


@router.get("/settings/auth")
def settings_auth_status(clients_settings_session: str | None = Cookie(None)):
    return {"authenticated": is_settings_authenticated(clients_settings_session)}


@router.delete("/settings/auth")
def settings_logout(response: Response, clients_settings_session: str | None = Cookie(None)):
    logout_settings(clients_settings_session)
    response.delete_cookie(COOKIE_NAME)
    return {"authenticated": False}

MANAGER_ORDER = [
    "Пашута М.С.", "Пашута М.С. (Ростов)", "Родина", "Родина Е.В. (Ростов)",
    "Новожилова М.", "Королева Светлана", "Ромащенко Екатерина", "Селянкина Татьяна",
    "Суркова Н.", "Трошина Лариса", "Шакулова Екатерина", "Антюфеева Яна",
    "Бабушкина Виктория", "Самойлова", "Андреева Дарья", "Гаина Татьяна",
    "Гордиенко", "Ермохина Ирина", "Кульченко Лилия", "Никишова Ольга",
    "Пименова Любовь", "Пирожкова Татьяна", "Стародубцева Полина", "Яицкая Ольга",
    "СОТРУДНИК АВИАТОРОВ", "СОТРУДНИК АХТУБИНСК", "СОТРУДНИК БАХТУРОВА",
    "СОТРУДНИК ЕВРОПА", "СОТРУДНИК ИДЕЯ", "СОТРУДНИК ПАРКХАУС",
    "СОТРУДНИК ПРИВОЗ", "СОТРУДНИК САНВЭЙ", "СОТРУДНИК СТРОЙГРАД",
    "СОТРУДНИК ТУЛАК", "СОТРУДНИК ЦИТРУС", "СОТРУДНИК ЦУМ",
    "Существующие сотрудники", "Клишко Ю.Н.", "МАРКЕТПЛЕЙСЫ", "Наш Китай",
    "Нет менеджера", "Дегтярев Алексей", "Дегтярева Оксана Александровна", "!!!", "<>",
]


@router.get("/health")
def health(db: Session = Depends(get_db)):
    db.execute(select(1))
    return {"status": "ok"}


def to_list_item(client: Client, last_import_at=None) -> ClientListItem:
    emails = sorted({email.email.strip() for email in client.emails if email.email and email.email.strip()})
    return ClientListItem(
        id=client.id,
        name=client.name,
        company=client.company,
        manager=client.manager,
        phone="\n".join(sorted({phone.phone for phone in client.phones})) or None,
        email="\n".join(emails) or None,
        trade_place=client.trade_places[0].place if client.trade_places else None,
        birth_date=client.birth_date,
        last_import_at=last_import_at,
        status=client.status.value,
    )


def apply_client_filters(
    query,
    *,
    search=None,
    phone_search=None,
    manager=None,
    company=None,
    price_type=None,
    buyer_type=None,
    counterparty_type=None,
    trade_place=None,
    has_email=None,
    has_phone=None,
    status=None,
    birth_day=None,
    birth_month=None,
):
    if search:
        term = f"%{search.lower()}%"
        query = query.where(func.lower(Client.name).like(term))
    if phone_search:
        query = query.where(Client.phones.any(Phone.phone == phone_search.strip()))
    if manager:
        include_empty_manager = "Нет менеджера" in manager
        selected_managers = [value for value in manager if value != "Нет менеджера"]
        manager_conditions = []
        if selected_managers:
            manager_conditions.append(Client.manager.in_(selected_managers))
        if include_empty_manager:
            manager_conditions.append(or_(Client.manager.is_(None), Client.manager == ""))
        query = query.where(or_(*manager_conditions))
    if company:
        query = query.where(Client.company == company)
    if price_type:
        query = query.where(Client.price_type.in_(price_type))
    if buyer_type:
        query = query.where(Client.buyer_type.in_(buyer_type))
    if counterparty_type:
        query = query.where(Client.counterparty_type.in_(counterparty_type))
    if trade_place:
        query = query.where(Client.trade_places.any(TradePlace.place == trade_place))
    if has_email is not None:
        # Email.email — обязательное нормализованное поле, поэтому достаточно
        # проверить непустое значение. Функции trim/length внутри relationship.any
        # на рабочем PostgreSQL приводили к ошибке выполнения запроса.
        has_filled_email = Client.emails.any(Email.email != "")
        query = query.where(has_filled_email if has_email else ~has_filled_email)
    if has_phone is not None:
        query = query.where(Client.phones.any() if has_phone else ~Client.phones.any())
    if status:
        query = query.where(Client.status == status)
    if birth_day:
        query = query.where(func.extract("day", Client.birth_date) == birth_day)
    if birth_month:
        query = query.where(func.extract("month", Client.birth_date) == birth_month)
    return query


@router.get("/clients", response_model=PagedClients)
def clients(
    db: Session = Depends(get_db),
    page: int = 1,
    page_size: str = "100",
    search: str | None = None,
    phone_search: str | None = None,
    manager: list[str] | None = Query(None),
    company: str | None = None,
    price_type: list[str] | None = Query(None),
    buyer_type: list[str] | None = Query(None),
    counterparty_type: list[str] | None = Query(None),
    trade_place: str | None = None,
    has_email: bool | None = None,
    has_phone: bool | None = None,
    status: str | None = None,
    birth_day: int | None = None,
    birth_month: int | None = None,
    sort: str = "name",
    order: str = "asc",
):
    page = max(page, 1)
    show_all = page_size == "all"
    try:
        parsed_page_size = 100 if show_all else int(page_size)
    except ValueError:
        parsed_page_size = 100
    parsed_page_size = min(max(parsed_page_size, 1), 500)
    filtered_ids = apply_client_filters(
        select(Client.id),
        search=search,
        phone_search=phone_search,
        manager=manager,
        company=company,
        price_type=price_type,
        buyer_type=buyer_type,
        counterparty_type=counterparty_type,
        trade_place=trade_place,
        has_email=has_email,
        has_phone=has_phone,
        status=status,
        birth_day=birth_day,
        birth_month=birth_month,
    ).distinct().subquery()
    total = db.scalar(select(func.count()).select_from(filtered_ids)) or 0
    sort_map = {
        "name": Client.name,
        "company": Client.company,
        "manager": Client.manager,
        "birth_date": Client.birth_date,
        "updated_at": Client.updated_at,
        "last_import": Import.imported_at,
    }
    sort_column = sort_map.get(sort, Client.name)
    order_by = sort_column.desc().nullslast() if order == "desc" else sort_column.asc().nullslast()
    availability_order = case((Client.status == ClientStatus.out_of_stock, 1), else_=0)
    stmt = (
        select(Client, Import.imported_at)
        .join(filtered_ids, filtered_ids.c.id == Client.id)
        .outerjoin(Import, Client.last_import_id == Import.id)
        .options(selectinload(Client.phones), selectinload(Client.emails), selectinload(Client.trade_places))
        .order_by(availability_order.asc(), order_by, Client.id.asc())
    )
    if show_all:
        page = 1
        response_page_size = total
    else:
        stmt = stmt.offset((page - 1) * parsed_page_size).limit(parsed_page_size)
        response_page_size = parsed_page_size
    items = [to_list_item(client, imported_at) for client, imported_at in db.execute(stmt).all()]
    return PagedClients(items=items, total=total, page=page, page_size=response_page_size)


@router.get("/clients-filter-options")
def client_filter_options(db: Session = Depends(get_db)):
    managers_from_db = db.scalars(
        select(Client.manager).where(Client.manager.is_not(None), Client.manager != "").distinct().order_by(Client.manager)
    ).all()
    manager_rank = {manager: index for index, manager in enumerate(MANAGER_ORDER)}
    managers = sorted(
        set(managers_from_db) | {"Нет менеджера"},
        key=lambda manager: (manager_rank.get(manager, len(MANAGER_ORDER)), manager.casefold()),
    )
    price_types = db.scalars(
        select(Client.price_type).where(Client.price_type.is_not(None), Client.price_type != "").distinct().order_by(Client.price_type)
    ).all()
    buyer_types = db.scalars(
        select(Client.buyer_type).where(Client.buyer_type.is_not(None), Client.buyer_type != "").distinct().order_by(Client.buyer_type)
    ).all()
    counterparty_types = db.scalars(
        select(Client.counterparty_type)
        .where(Client.counterparty_type.is_not(None), Client.counterparty_type != "")
        .distinct()
        .order_by(Client.counterparty_type)
    ).all()
    return {
        "managers": managers,
        "price_types": price_types,
        "buyer_types": buyer_types,
        "counterparty_types": counterparty_types,
    }


@router.get("/clients/{client_id}", response_model=ClientDetail)
def client_detail(client_id: int, db: Session = Depends(get_db)):
    client = db.scalar(
        select(Client)
        .where(Client.id == client_id)
        .options(selectinload(Client.phones), selectinload(Client.emails), selectinload(Client.trade_places))
    )
    if client is None:
        raise HTTPException(status_code=404, detail="Клиент не найден")
    first_import = db.get(Import, client.first_import_id) if client.first_import_id else None
    last_import = db.get(Import, client.last_import_id) if client.last_import_id else None
    base = to_list_item(client, last_import.imported_at if last_import else None).model_dump()
    base.update(
        price_type=client.price_type,
        director=client.director,
        contact_person=client.contact_person,
        raw_common_phones=client.raw_common_phones,
        raw_sms_phones=client.raw_sms_phones,
        raw_email=client.raw_email,
        raw_email_source_known=client.raw_email_source_known,
        client_source=client.client_source,
        last_purchase_date=client.last_purchase_date,
        buyer_type=client.buyer_type,
        counterparty_type=client.counterparty_type,
        created_at=client.created_at,
        updated_at=client.updated_at,
        first_import_at=first_import.imported_at if first_import else None,
        last_import_file=last_import.file_name if last_import else None,
        phones=[{"phone": phone.phone, "type": phone.type.value} for phone in client.phones],
        emails=[email.email for email in client.emails],
        trade_places=[place.place for place in client.trade_places],
    )
    return ClientDetail(**base)


@router.post("/imports")
async def upload_import(background_tasks: BackgroundTasks, files: list[UploadFile] = File(...), _: None = Depends(require_settings_auth)):
    started = monotonic()
    import_logger.info("Начало загрузки API: файлов=%s", len(files))
    payload = [(file.filename or "import.xlsx", await file.read()) for file in files]
    import_logger.info(
        "Файлы получены API: файлов=%s, байт=%s, длительность чтения upload=%.3f сек",
        len(payload), sum(len(content) for _, content in payload), monotonic() - started,
    )
    task_id = create_import_task(payload)
    background_tasks.add_task(run_import_task, task_id, payload)
    import_logger.info("Задача создана: task_id=%s, ответ подготовлен за %.3f сек", task_id, monotonic() - started)
    return {"status": "accepted", "task_id": task_id}


@router.get("/imports/tasks/{task_id}")
def import_task(task_id: str, _: None = Depends(require_settings_auth)):
    task = get_import_task(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="Задача импорта не найдена")
    return task


@router.get("/imports")
def imports(db: Session = Depends(get_db), _: None = Depends(require_settings_auth)):
    return db.scalars(select(Import).order_by(Import.imported_at.desc()).limit(200)).all()


@router.get("/imports/{import_id}/issues")
def import_issues(import_id: int, db: Session = Depends(get_db), _: None = Depends(require_settings_auth)):
    return db.scalars(select(ImportIssue).where(ImportIssue.import_id == import_id).order_by(ImportIssue.id)).all()




@router.get("/logs")
def logs(source: str | None = None, db: Session = Depends(get_db), _: None = Depends(require_settings_auth)):
    imports = db.scalars(
        select(Import).where(~Import.log_hidden).order_by(Import.id.desc()).limit(100)
    ).all()
    import_ids = [item.id for item in imports]
    first_errors: dict[int, str] = {}
    if import_ids:
        errors = db.execute(
            select(ImportIssue.import_id, ImportIssue.message)
            .where(ImportIssue.import_id.in_(import_ids), ImportIssue.level == "error")
            .order_by(ImportIssue.id)
        ).all()
        for import_id, message in errors:
            first_errors.setdefault(import_id, message)
    items = [
        {
            "id": f"import-{item.id}",
            "created_at": item.imported_at,
            "source": "Импорт",
            "level": "error" if item.error_count else "info",
            "process": item.file_name,
            "row_number": None,
            "message": (
                f"Найдено строк: {item.rows_count}. Загружено: {item.added_count}. "
                f"Обновлено: {item.updated_count}. Пропущено: {item.skipped_count}. Ошибок: {item.error_count}."
                + (f" Ошибка: {first_errors[item.id]}" if item.id in first_errors else "")
            ),
        }
        for item in imports
    ]
    ftp_events = db.scalars(select(FtpImportEvent).order_by(FtpImportEvent.id.desc()).limit(100)).all()
    items.extend({
        "id": f"ftp-{item.id}", "created_at": item.created_at, "source": "FTP",
        "level": "error" if item.status == "Ошибка" else "warning" if item.status == "С предупреждениями" else "info",
        "process": item.file_name,
        "row_number": None,
        "message": (
            f"Размер: {item.file_size} байт. Загружено: {item.added_count}. Обновлено: {item.updated_count}. "
            f"Пропущено: {item.skipped_count}. Длительность: {item.duration_seconds} сек. Статус: {item.status}."
            + (
                f" {'Предупреждение' if item.status == 'С предупреждениями' else 'Ошибка'}: {item.error}"
                if item.error else ""
            )
        ),
    } for item in ftp_events)
    if source:
        items = [item for item in items if item["source"].lower() == source.lower()]
    items.sort(key=lambda item: item["created_at"], reverse=True)
    return items[:100]


@router.delete("/logs")
def delete_logs(db: Session = Depends(get_db), _: None = Depends(require_settings_auth)):
    # Import нельзя удалять физически: на него ссылаются карточки клиентов.
    # Поэтому служебные записи скрываем, а собственно события удаляем полностью.
    hidden = db.execute(update(Import).where(~Import.log_hidden).values(log_hidden=True)).rowcount or 0
    db.execute(delete(ImportIssue))
    db.execute(delete(AuditLog))
    db.execute(delete(FtpImportEvent))
    db.commit()
    return {"deleted": hidden}


@router.get("/ftp/settings")
def ftp_settings(_: None = Depends(require_settings_auth)):
    settings = get_ftp_settings().model_dump(exclude={"password"})
    settings["password"] = ""
    settings["password_configured"] = bool(get_ftp_settings().password)
    return settings


@router.put("/ftp/settings")
def update_ftp_settings(payload: dict, _: None = Depends(require_settings_auth)):
    settings = save_ftp_settings(payload)
    refresh_ftp_schedule()
    return {**settings.model_dump(exclude={"password"}), "password": "", "password_configured": bool(settings.password)}


@router.get("/ftp/status")
def ftp_status(_: None = Depends(require_settings_auth)):
    return get_ftp_status()


@router.post("/ftp/test")
def ftp_test(_: None = Depends(require_settings_auth)):
    try:
        return test_connection()
    except Exception as error:
        raise HTTPException(status_code=400, detail=f"Не удалось подключиться к FTP: {error}") from error


@router.post("/ftp/run")
def ftp_run(background_tasks: BackgroundTasks, _: None = Depends(require_settings_auth)):
    if get_ftp_status()["running"]:
        raise HTTPException(status_code=409, detail="Автозагрузка уже выполняется")
    mark_ftp_pending()
    background_tasks.add_task(run_ftp_import, retry_when_empty=False)
    return {"status": "accepted"}

@router.post("/clients/bulk")
def bulk_update(payload: BulkUpdate, db: Session = Depends(get_db)):
    clients_to_update = db.scalars(
        select(Client).where(Client.id.in_(payload.ids)).options(
            selectinload(Client.phones), selectinload(Client.emails), selectinload(Client.trade_places)
        )
    ).all()
    for client in clients_to_update:
        previous_signature = payload_signature(client)
        if payload.manager is not None:
            client.manager = payload.manager
        if payload.price_type is not None:
            client.price_type = payload.price_type
        if payload.status is not None:
            client.status = payload.status
        if payload_signature(client) != previous_signature:
            record_client_change(db, client)
        db.add(AuditLog(client_id=client.id, action="bulk_update", payload=payload.model_dump_json(exclude_none=True)))
    db.commit()
    return {"updated": len(clients_to_update)}


@router.delete("/clients")
def bulk_delete(
    ids: str | None = None,
    delete_all: bool = Query(False, alias="all"),
    db: Session = Depends(get_db),
    clients_settings_session: str | None = Cookie(None),
):
    if delete_all:
        if not is_settings_authenticated(clients_settings_session):
            raise HTTPException(status_code=401, detail="Введите пароль для доступа к настройкам")
        # Связанные телефоны, email и места торговли удаляются на уровне БД
        # благодаря внешним ключам с ON DELETE CASCADE.
        # Tombstone создаётся одним INSERT SELECT до удаления клиентов.
        db.execute(insert(ClientChange).from_select(
            [ClientChange.client_id, ClientChange.operation],
            select(Client.id, literal("delete")),
        ))
        deleted = db.execute(delete(Client)).rowcount or 0
        db.add(AuditLog(action="delete_all_clients", payload=f"Удалено клиентов: {deleted}"))
        db.commit()
        return {"deleted": deleted}
    if not ids:
        raise HTTPException(status_code=400, detail="Не переданы строки для удаления")
    id_list = [int(value) for value in ids.split(",") if value.strip()]
    clients_to_delete = db.scalars(select(Client).where(Client.id.in_(id_list))).all()
    for client in clients_to_delete:
        record_client_change(db, client, operation="delete")
        db.delete(client)
    db.add(AuditLog(action="bulk_delete", payload=",".join(map(str, id_list))))
    db.commit()
    return {"deleted": len(clients_to_delete)}


@router.get("/clients-export.xlsx")
def export_clients(
    db: Session = Depends(get_db),
    search: str | None = None,
    phone_search: str | None = None,
    manager: list[str] | None = Query(None),
    company: str | None = None,
    price_type: list[str] | None = Query(None),
    buyer_type: list[str] | None = Query(None),
    counterparty_type: list[str] | None = Query(None),
    trade_place: str | None = None,
    has_email: bool | None = None,
    has_phone: bool | None = None,
    status: str | None = None,
    birth_day: int | None = None,
    birth_month: int | None = None,
    columns: list[str] | None = Query(None),
):
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet("clients")
    selected_columns = [value for value in (columns or list(EXPORT_COLUMNS)) if value in EXPORT_COLUMNS]
    if not selected_columns:
        selected_columns = ["name"]
    headers = [EXPORT_COLUMNS[value][0] for value in selected_columns]
    for column, header in enumerate(headers):
        worksheet.write(0, column, header)
    filtered_ids = apply_client_filters(
        select(Client.id),
        search=search,
        phone_search=phone_search,
        manager=manager,
        company=company,
        price_type=price_type,
        buyer_type=buyer_type,
        counterparty_type=counterparty_type,
        trade_place=trade_place,
        has_email=has_email,
        has_phone=has_phone,
        status=status,
        birth_day=birth_day,
        birth_month=birth_month,
    ).distinct().subquery()
    stmt = (
        select(Client)
        .join(filtered_ids, filtered_ids.c.id == Client.id)
        .options(selectinload(Client.phones), selectinload(Client.emails), selectinload(Client.trade_places))
        .order_by(case((Client.status == ClientStatus.out_of_stock, 1), else_=0), Client.name, Client.id)
    )
    for row_number, client in enumerate(db.scalars(stmt), start=1):
        worksheet.write_row(row_number, 0, [EXPORT_COLUMNS[value][1](client) for value in selected_columns])
    workbook.close()
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clients.xlsx"},
    )


def _email_report_xlsx(db: Session, report: dict) -> tuple[bytes, int]:
    """Формирует один email-файл с дедупликацией клиентов по наименованию."""
    filtered_ids = apply_client_filters(
        select(Client.id), manager=report.get("managers") or None,
        price_type=report.get("price_types") or None,
        buyer_type=report.get("buyer_types") or None,
    ).distinct().subquery()
    clients = db.scalars(
        select(Client).join(filtered_ids, filtered_ids.c.id == Client.id)
        .where(or_(
            Client.raw_email != "",
            (Client.raw_email_source_known.is_(False) & Client.emails.any(Email.email != "")),
        ))
        .options(selectinload(Client.emails))
        .order_by(Client.name, Client.id)
    ).all()
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output, {"in_memory": True})
    worksheet = workbook.add_worksheet("Email")
    worksheet.write_row(0, 0, ["Наименование", "Email"])
    excluded_emails = {
        value.casefold() for value in db.scalars(select(EmailReportExclusion.email)).all()
    }
    grouped: dict[str, tuple[str, set[str]]] = {}
    for client in clients:
        display_name = (client.name or "").strip()
        if _email_report_excludes_name(report, display_name):
            continue
        name_key = display_name.casefold()
        if not name_key:
            continue
        if name_key not in grouped:
            grouped[name_key] = (display_name, set())
        # До первого повторного импорта источник старых email определить нельзя.
        # Для таких legacy-записей используем прежний список, чтобы отчёт не был
        # пустым; после импорта берём строго исходное поле карточки.
        emails = (
            extract_emails(client.raw_email)
            if client.raw_email_source_known
            else [item.email.strip() for item in client.emails if item.email.strip()]
        )
        grouped[name_key][1].update(email for email in emails if email.casefold() not in excluded_emails)
    row = 1
    for display_name, emails in grouped.values():
        for email in sorted(emails):
            worksheet.write_row(row, 0, [display_name, email])
            row += 1
    worksheet.set_column(0, 0, 42)
    worksheet.set_column(1, 1, 38)
    workbook.close()
    return output.getvalue(), row - 1


def _email_report_excludes_name(report: dict, name: str) -> bool:
    """Исключает организации из розничного email-файла по аббревиатуре."""
    report_name = str(report.get("name") or "").strip()
    return (
        report_name.casefold() == RETAIL_EMAIL_REPORT_NAME.casefold()
        and RETAIL_EMAIL_REPORT_EXCLUDED_NAME_PATTERN.search(name) is not None
    )


def _emails_from_exclusion_xlsx(content: bytes) -> set[str]:
    """Читает и нормализует адреса из колонки Email первой страницы XLSX."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as error:
        raise HTTPException(status_code=422, detail="Не удалось прочитать XLSX-файл") from error
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    headers = next(rows, ())
    email_column = next(
        (index for index, value in enumerate(headers) if str(value or "").strip().casefold() == "email"),
        None,
    )
    if email_column is None:
        raise HTTPException(status_code=422, detail="В XLSX-файле не найдена колонка Email")
    emails: set[str] = set()
    for row in rows:
        if email_column >= len(row):
            continue
        emails.update(email.casefold() for email in extract_emails(str(row[email_column] or "")))
    return emails


def _validate_email_exclusion_category(category: str) -> None:
    if category not in EMAIL_EXCLUSION_CATEGORIES:
        raise HTTPException(status_code=404, detail="Неизвестный список исключений")


@router.get("/reports/email-update/exclusions/{category}")
def email_exclusions(category: str, db: Session = Depends(get_db)):
    _validate_email_exclusion_category(category)
    return {
        "category": category,
        "emails": db.scalars(
            select(EmailReportExclusion.email)
            .where(EmailReportExclusion.category == category)
            .order_by(EmailReportExclusion.email)
        ).all(),
    }


@router.post("/reports/email-update/exclusions/{category}")
async def upload_email_exclusions(category: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    _validate_email_exclusion_category(category)
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Загрузите файл в формате XLSX")
    emails = _emails_from_exclusion_xlsx(await file.read())
    existing_emails = set(db.scalars(
        select(EmailReportExclusion.email).where(EmailReportExclusion.category == category)
    ).all())
    new_emails = emails - existing_emails
    if new_emails:
        db.execute(insert(EmailReportExclusion), [
            {"category": category, "email": email} for email in sorted(new_emails)
        ])
    db.commit()
    updated_emails = sorted(existing_emails | emails)
    return {
        "category": category,
        "emails": updated_emails,
        "count": len(updated_emails),
        "added_count": len(new_emails),
    }


def _email_report_filename(name: str, row_count: int, index: int, used_names: set[str]) -> str:
    """Добавляет количество строк и гарантирует уникальность имени в ZIP."""
    safe_name = "".join(character for character in name if character not in '\\/:*?"<>|').strip() or f"Отчет {index}"
    filename = f"{safe_name}. {row_count}.xlsx"
    if filename.casefold() in used_names:
        filename = f"{safe_name} ({index}). {row_count}.xlsx"
    used_names.add(filename.casefold())
    return filename


@router.get("/reports/email-update/config")
def email_update_config(db: Session = Depends(get_db)):
    managers = db.scalars(
        select(Client.manager).where(Client.manager.is_not(None), Client.manager != "").distinct().order_by(Client.manager)
    ).all()
    price_types = db.scalars(
        select(Client.price_type).where(Client.price_type.is_not(None), Client.price_type != "").distinct().order_by(Client.price_type)
    ).all()
    buyer_types = db.scalars(
        select(Client.buyer_type).where(Client.buyer_type.is_not(None), Client.buyer_type != "").distinct().order_by(Client.buyer_type)
    ).all()
    return {"reports": DEFAULT_EMAIL_REPORTS, "managers": managers, "price_types": price_types, "buyer_types": buyer_types}


@router.post("/reports/email-update.zip")
def email_update_report(payload: dict, db: Session = Depends(get_db)):
    reports = payload.get("reports")
    if not isinstance(reports, list) or not reports:
        raise HTTPException(status_code=422, detail="Добавьте хотя бы один файл отчёта")
    archive = BytesIO()
    used_names: set[str] = set()
    with ZipFile(archive, "w", ZIP_DEFLATED) as zipped:
        for index, report in enumerate(reports, start=1):
            name = str(report.get("name") or f"Отчет {index}").strip()
            content, row_count = _email_report_xlsx(db, report)
            filename = _email_report_filename(name, row_count, index, used_names)
            zipped.writestr(filename, content)
    archive.seek(0)
    return StreamingResponse(
        archive, media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=email-update.zip"},
    )
