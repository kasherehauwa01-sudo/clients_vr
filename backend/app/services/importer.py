from dataclasses import dataclass, field
import logging
from logging.handlers import RotatingFileHandler
import os
from pathlib import Path
import re
from tempfile import TemporaryDirectory
from time import monotonic
from typing import Callable
from zipfile import BadZipFile
import magic
import xlrd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session
from app.models.entities import AuditLog, Client, Email, Import, ImportIssue, Phone, PhoneType, TradePlace
from app.services.client_changes import payload_signature, record_change_if_needed
from app.services.normalization import clean_multiline_text, clean_text, extract_emails, extract_phones, parse_date, repair_legacy_excel_text, split_values

logger = logging.getLogger(__name__)
import_logger = logging.getLogger("clients.import")
if not import_logger.handlers:
    import_log_path = Path(os.getenv("CLIENTS_IMPORT_LOG", "/tmp/clients-import.log"))
    import_log_path.parent.mkdir(parents=True, exist_ok=True)
    handler = RotatingFileHandler(import_log_path, maxBytes=10 * 1024 * 1024, backupCount=5, encoding="utf-8")
    handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    import_logger.addHandler(handler)
    import_logger.setLevel(logging.INFO)
    import_logger.propagate = False

ProgressCallback = Callable[[str, int, int], None]

COLUMN_ORDER = [
    "name",
    "price_type",
    "manager",
    "birth_date",
    "emails",
    "common_phones",
    "trade_places",
    "sms_phones",
    "director",
    "company",
    "contact_person",
    "client_source",
    "last_purchase_date",
    "buyer_type",
    "counterparty_type",
]

FIELD_LABELS = {
    "name": "Наименование",
    "price_type": "Тип цены",
    "manager": "Менеджер",
    "birth_date": "Дата рождения",
    "emails": "Email",
    "common_phones": "Телефоны прочие",
    "trade_places": "Места торговли",
    "sms_phones": "Телефоны для СМС и рассылки",
    "director": "Руководитель",
    "contact_person": "Контактное лицо",
    "company": "Фирма",
    "client_source": "Источник клиента",
    "last_purchase_date": "Дата последней покупки",
    "buyer_type": "Вид покупателя",
    "counterparty_type": "Вид контрагента",
}

HEADER_ALIASES = {
    "наименование": "name",
    "название": "name",
    "клиент": "name",
    "типцены": "price_type",
    "ценатип": "price_type",
    "менеджер": "manager",
    "датарождения": "birth_date",
    "др": "birth_date",
    "email": "emails",
    "emails": "emails",
    "почта": "emails",
    "телефоныпрочие": "common_phones",
    "прочиетелефоны": "common_phones",
    "телефон": "common_phones",
    "телефоны": "common_phones",
    "местаторговли": "trade_places",
    "местоторговли": "trade_places",
    "адрес": "trade_places",
    "телефоныдлясмсирассылки": "sms_phones",
    "телефондлясмсирассылки": "sms_phones",
    "телефоныдлярассылки": "sms_phones",
    "телефонсмс": "sms_phones",
    "руководитель": "director",
    "контактноелицо": "contact_person",
    "контакт": "contact_person",
    "фирма": "company",
    "компания": "company",
    "источникклиента": "client_source",
    "источник": "client_source",
    "датапоследнейпокупки": "last_purchase_date",
    "последняяпокупка": "last_purchase_date",
    "видпокупателя": "buyer_type",
    "типпокупателя": "buyer_type",
    "видконтрагента": "counterparty_type",
    "типконтрагента": "counterparty_type",
}


@dataclass
class WorkbookRows:
    rows: list[list[object]]
    file_format: str
    sheet_count: int
    sheet_name: str
    repaired_cells: int = 0
    parser: str = "unknown"


@dataclass
class ParsedRows:
    rows: list[dict[str, object]]
    logs: list[str] = field(default_factory=list)
    total_rows: int = 0
    read_rows: int = 0


class ImportSummary(BaseModel):
    files: int = 0
    rows: int = 0
    read: int = 0
    added: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    duplicates: int = 0
    logs: list[str] = Field(default_factory=list)


def _normalize_header(value: object) -> str:
    text = clean_text(value) or ""
    return re.sub(r"[^0-9a-zа-яё]+", "", text.lower())


def _row_has_data(row: list[object]) -> bool:
    return any(clean_text(value) for value in row)


def _map_by_positions(row: list[object]) -> dict[str, object]:
    return {field: row[index] if index < len(row) else None for index, field in enumerate(COLUMN_ORDER)}


def _detect_header(rows: list[list[object]]) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_mapping: dict[int, str] = {}
    for index, row in enumerate(rows[:20]):
        mapping = {column_index: HEADER_ALIASES[key] for column_index, value in enumerate(row) if (key := _normalize_header(value)) in HEADER_ALIASES}
        if len(mapping) > len(best_mapping):
            best_index, best_mapping = index, mapping
    return (best_index, best_mapping) if len(best_mapping) >= 2 else (None, {})


OLE_SIGNATURE = bytes.fromhex("D0 CF 11 E0 A1 B1 1A E1")
ZIP_SIGNATURE = b"PK"


def _rows_from_xlsx(path: Path) -> WorkbookRows:
    # Передаём поток, чтобы openpyxl также корректно читал XLSX с ошибочным
    # расширением .xls: формат уже надёжно определён по ZIP-сигнатуре.
    with path.open("rb") as source:
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheet_names = list(workbook.sheetnames)
        logger.info(
            "XLSX открыт через openpyxl: листов=%s, названия=%s, строк=%s, столбцов=%s",
            len(sheet_names), sheet_names, sheet.max_row, sheet.max_column,
        )
        return WorkbookRows(rows=rows, file_format="xlsx", sheet_count=len(sheet_names), sheet_name=sheet.title, parser="openpyxl")


def _rows_from_xls_xlrd(path: Path) -> WorkbookRows:
    book = xlrd.open_workbook(str(path))
    sheet = book.sheet_by_index(0)
    logger.info(
        "XLS открыт через xlrd: листов=%s, названия=%s, строк=%s, столбцов=%s",
        book.nsheets, book.sheet_names(), sheet.nrows, sheet.ncols,
    )
    rows: list[list[object]] = []
    repaired_cells = 0
    for row_index in range(sheet.nrows):
        values: list[object] = []
        for column_index in range(sheet.ncols):
            cell = sheet.cell(row_index, column_index)
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    values.append(xlrd.xldate.xldate_as_datetime(cell.value, book.datemode))
                except Exception:
                    values.append(cell.value)
            else:
                value = repair_legacy_excel_text(cell.value)
                repaired_cells += int(value != cell.value)
                values.append(value)
        rows.append(values)
    return WorkbookRows(rows=rows, file_format="xls", sheet_count=book.nsheets, sheet_name=sheet.name, repaired_cells=repaired_cells, parser="xlrd")


def _detect_workbook_format(content: bytes) -> str:
    if content.startswith(OLE_SIGNATURE):
        return "xls"
    if content.startswith(ZIP_SIGNATURE):
        return "xlsx"
    signature = content[:8].hex(" ").upper() or "пустой файл"
    raise ValueError(
        f"Не удалось определить формат Excel по сигнатуре ({signature}). "
        "Загрузите настоящий XLS (Excel 97–2003) или XLSX."
    )


def _read_workbook(filename: str, content: bytes) -> WorkbookRows:
    extension = Path(filename).suffix.lower() or "без расширения"
    with TemporaryDirectory(prefix="clients-import-") as directory:
        path = Path(directory) / f"upload{extension if extension.startswith('.') else '.bin'}"
        path.write_bytes(content)
        saved_content = path.read_bytes()
        if saved_content != content:
            raise IOError("Загруженный файл изменился при сохранении на диск")
        mime_type = magic.from_file(str(path), mime=True)
        try:
            file_format = _detect_workbook_format(saved_content)
        except ValueError:
            logger.error(
                "Неподдерживаемый файл: путь=%s; размер=%s байт; первые 32 байта=%s; MIME=%s; расширение=%s",
                path.resolve(), len(saved_content), saved_content[:32].hex(" ").upper(), mime_type, extension,
            )
            raise
        logger.info(
            "Диагностика файла: путь=%s; размер=%s байт; первые 32 байта=%s; MIME=%s; расширение=%s; формат по сигнатуре=%s; сохранение проверено",
            path.resolve(), len(saved_content), saved_content[:32].hex(" ").upper(), mime_type, extension, file_format,
        )
        try:
            if file_format == "xls":
                return _rows_from_xls_xlrd(path)
            return _rows_from_xlsx(path)
        except Exception:
            # logger.exception сохраняет тип ошибки, сообщение и полный traceback.
            logger.exception("Ошибка чтения файла %s парсером %s", path, "xlrd" if file_format == "xls" else "openpyxl")
            raise


def _describe_import_error(filename: str, error: Exception) -> str:
    """Return an actionable, user-facing reason without hiding the original error."""
    details = str(error).strip() or error.__class__.__name__
    lower_details = details.lower()
    if isinstance(error, (BadZipFile, InvalidFileException)) or "not a zip file" in lower_details:
        advice = "Файл не является корректным XLSX. Пересохраните его в Excel как .xlsx или загрузите исходный .xls."
    elif isinstance(error, xlrd.biffh.XLRDError):
        advice = "Не удалось прочитать формат XLS. Проверьте, что файл не поврежден и действительно сохранен как Excel 97–2003 (.xls)."
    elif isinstance(error, AssertionError):
        advice = "Внутренняя структура XLS повреждена. Откройте файл в Excel или LibreOffice, выберите «Сохранить как» и сохраните заново в формате .xlsx."
    elif "поддерживаются только" in lower_details:
        advice = "Допустимы только файлы с расширением .xls или .xlsx."
    else:
        advice = "Проверьте формат данных и строку файла, указанную в сообщении, затем повторите загрузку."
    return f"Ошибка файла «{filename}»: {details}. Что исправить: {advice}"


def _format_preview_value(value: object) -> str:
    text = clean_text(value)
    return text if text is not None else "NULL"


def _add_header_logs(logs: list[str], raw_rows: list[list[object]], header_index: int | None, mapping: dict[int, str]) -> None:
    if header_index is None:
        logs.append("Строка заголовков не найдена. Используется ожидаемый порядок столбцов.")
        for index, field_name in enumerate(COLUMN_ORDER):
            logs.append(f"Найдена колонка: {FIELD_LABELS[field_name]} -> column {index} (по порядку)")
        return
    header_row = raw_rows[header_index]
    logs.append("Заголовки:")
    for value in header_row:
        if clean_text(value):
            logs.append(str(clean_text(value)))
    reversed_mapping = {field_name: column_index for column_index, field_name in mapping.items()}
    for field_name in COLUMN_ORDER:
        label = FIELD_LABELS[field_name]
        column_index = reversed_mapping.get(field_name)
        if column_index is None:
            logs.append(f'Колонка "{label}" отсутствует')
            logs.append("Используется NULL")
        else:
            logs.append(f"Найдена колонка: {label} -> column {column_index}")


def _read_rows(filename: str, content: bytes) -> ParsedRows:
    workbook_rows = _read_workbook(filename, content)
    raw_rows = workbook_rows.rows
    logs = [
        f"Файл: {filename}",
        f"Формат: {workbook_rows.file_format}",
        f"Парсер: {workbook_rows.parser}",
        f"Листов: {workbook_rows.sheet_count}",
        f"Лист: {workbook_rows.sheet_name}",
        f"Строк: {len(raw_rows)}",
    ]
    if workbook_rows.repaired_cells:
        logs.append(f"Исправлена кодировка ячеек: {workbook_rows.repaired_cells}")
    non_empty_rows = [row for row in raw_rows if _row_has_data(row)]
    if not non_empty_rows:
        logs.append("Файл не содержит строк с данными.")
        return ParsedRows(rows=[], logs=logs, total_rows=len(raw_rows), read_rows=0)

    header_index, mapping = _detect_header(non_empty_rows)
    _add_header_logs(logs, non_empty_rows, header_index, mapping)
    start_row = header_index + 2 if header_index is not None else 1
    logs.append(f"Импорт начинается со строки: {start_row}")
    data_rows = non_empty_rows[header_index + 1 :] if header_index is not None else non_empty_rows

    result: list[dict[str, object]] = []
    for offset, row in enumerate(data_rows, start=start_row):
        if not _row_has_data(row):
            logs.append(f"Строка {offset}. Причина: пустая строка")
            continue
        mapped = {field: row[index] if index < len(row) else None for index, field in mapping.items()} if mapping else _map_by_positions(row)
        mapped["_row_number"] = offset
        result.append(mapped)

    for row in result[:10]:
        logs.extend(
            [
                f"Строка {row.get('_row_number')}",
                f"Наименование: {_format_preview_value(row.get('name'))}",
                f"Email: {_format_preview_value(row.get('emails'))}",
                f"Телефон: {_format_preview_value(row.get('sms_phones') or row.get('common_phones'))}",
            ]
        )
    return ParsedRows(rows=result, logs=logs, total_rows=len(raw_rows), read_rows=len(result))


def _company_matches(client: Client, company: str | None) -> bool:
    # В XLS один и тот же контактный телефон/email может относиться к разным фирмам.
    # Поэтому не склеиваем разные строки клиента, если в файле явно указана другая фирма.
    existing_company = clean_text(client.company)
    return not company or not existing_company or existing_company == company


def _find_client(db: Session, row: dict, sms: list[str], phones: list[str], emails: list[str]) -> tuple[Client | None, bool]:
    name, company = clean_text(row.get("name")), clean_text(row.get("company"))
    if name and company:
        found = db.scalars(select(Client).where(Client.name == name, Client.company == company)).all()
        if found:
            return found[0], len(found) > 1
    for phone_set, type_filter in ((sms, PhoneType.sms), (phones + sms, None)):
        if phone_set:
            query = select(Client).join(Phone).where(Phone.phone.in_(phone_set))
            if type_filter:
                query = query.where(Phone.type == type_filter)
            found = [client for client in db.scalars(query).unique().all() if _company_matches(client, company)]
            if found:
                return found[0], len(found) > 1
    if emails:
        found = [
            client
            for client in db.scalars(select(Client).join(Email).where(Email.email.in_(emails))).unique().all()
            if _company_matches(client, company)
        ]
        if found:
            return found[0], len(found) > 1
    return None, False


def _apply_client_data(client: Client, data: dict[str, object]) -> None:
    for key, value in data.items():
        # Пустые значения из очередной строки XLS не должны затирать уже загруженные данные.
        # Для исходного Email пустота значима: она означает, что поле карточки
        # проверено и действительно не содержит адреса.
        if key in {"last_import_id", "raw_email", "raw_email_source_known"} or value not in (None, ""):
            setattr(client, key, value)


def _sync_children(client: Client, emails: list[str], sms: list[str], common: list[str], places: list[str]) -> None:
    existing_emails = {email.email for email in client.emails}
    for email in dict.fromkeys(emails):
        if email not in existing_emails:
            client.emails.append(Email(email=email))
            existing_emails.add(email)
    existing_phones = {(phone.phone, phone.type) for phone in client.phones}
    for phone, phone_type in [(value, PhoneType.sms) for value in sms] + [(value, PhoneType.common) for value in common]:
        if (phone, phone_type) not in existing_phones:
            normalized = normalize_phone(phone)
            client.phones.append(Phone(phone=phone, normalized_phone=normalized[-10:] if normalized else None, type=phone_type))
            existing_phones.add((phone, phone_type))
    existing_places = {place.place for place in client.trade_places}
    for place in dict.fromkeys(places):
        if place not in existing_places:
            client.trade_places.append(TradePlace(place=place))
            existing_places.add(place)


def _fallback_name(row: dict, emails: list[str], sms: list[str], common: list[str]) -> str:
    return (
        clean_text(row.get("name"))
        or clean_text(row.get("company"))
        or clean_text(row.get("contact_person"))
        or clean_text(row.get("director"))
        or (emails[0] if emails else None)
        or (sms[0] if sms else None)
        or (common[0] if common else None)
        or f"Клиент из строки {row.get('_row_number', '')}".strip()
    )


def _log_issue(db: Session, import_id: int, message: str, *, row_number: int | None = None, level: str = "info") -> None:
    db.add(ImportIssue(import_id=import_id, row_number=row_number, level=level, message=message))


def import_files(db: Session, files: list[tuple[str, bytes]], progress: ProgressCallback | None = None) -> ImportSummary:
    import_started = monotonic()
    import_logger.info("Начало загрузки: файлов=%s", len(files))
    total = ImportSummary(files=len(files))
    for filename, content in files:
        file_started = monotonic()
        import_logger.info("Файл получен: имя=%s, размер=%s байт", filename, len(content))
        if progress:
            progress("Чтение…", 0, 0)
        imp = Import(file_name=filename, rows_count=0, added_count=0, updated_count=0, skipped_count=0, error_count=0)
        db.add(imp)
        db.flush()
        try:
            parsed = _read_rows(filename, content)
            import_logger.info(
                "Файл открыт и прочитан: имя=%s, найдено=%s, строк для обработки=%s, длительность=%.3f сек",
                filename, parsed.total_rows, parsed.read_rows, monotonic() - file_started,
            )
            if progress:
                progress("Обработка…", 0, parsed.read_rows)
            total.logs.extend(parsed.logs)
            for message in parsed.logs:
                _log_issue(db, imp.id, message)
            imp.rows_count = parsed.read_rows
            total.rows += parsed.total_rows
            total.read += parsed.read_rows
            processing_started = monotonic()
            for processed, row in enumerate(parsed.rows, start=1):
                row_number = int(row.get("_row_number") or 0) or None
                try:
                    created = False
                    duplicate = False
                    client_id: int | None = None
                    action = "client_updated"
                    with db.begin_nested():
                        direct_emails = extract_emails(row.get("emails"))
                        emails = list(
                            dict.fromkeys(
                                direct_emails
                                + extract_emails(row.get("trade_places"))
                                + extract_emails(row.get("director"))
                                + extract_emails(row.get("contact_person"))
                            )
                        )
                        sms = extract_phones(row.get("sms_phones"))
                        common = list(
                            dict.fromkeys(
                                extract_phones(row.get("common_phones"))
                                + extract_phones(row.get("trade_places"))
                                + extract_phones(row.get("director"))
                                + extract_phones(row.get("contact_person"))
                            )
                        )
                        places = [place for place in (clean_text(value) for value in split_values(row.get("trade_places"))) if place]
                        name = _fallback_name(row, emails, sms, common)
                        client, duplicate = _find_client(db, row | {"name": name}, sms, common, emails)
                        previous_signature = payload_signature(client) if client else None
                        data = dict(
                            name=name,
                            company=clean_text(row.get("company")),
                            price_type=clean_text(row.get("price_type")),
                            manager=clean_text(row.get("manager")),
                            birth_date=parse_date(row.get("birth_date")),
                            director=clean_text(row.get("director")),
                            contact_person=clean_text(row.get("contact_person")),
                            raw_common_phones=clean_multiline_text(row.get("common_phones")),
                            raw_sms_phones=clean_multiline_text(row.get("sms_phones")),
                            raw_email="\n".join(direct_emails) or None,
                            raw_email_source_known=True,
                            client_source=clean_text(row.get("client_source")),
                            last_purchase_date=parse_date(row.get("last_purchase_date")),
                            buyer_type=clean_text(row.get("buyer_type")),
                            counterparty_type=clean_text(row.get("counterparty_type")),
                            last_import_id=imp.id,
                        )
                        if client:
                            _apply_client_data(client, data)
                        else:
                            client = Client(**data, first_import_id=imp.id)
                            db.add(client)
                            created = True
                            action = "client_created"
                        _sync_children(client, emails, sms, common, places)
                        db.flush()
                        client_id = client.id
                        record_change_if_needed(db, client, previous_signature)
                    if duplicate:
                        total.duplicates += 1
                        _log_issue(db, imp.id, "Найдено несколько совпадений клиента", row_number=row_number, level="warning")
                    if created:
                        imp.added_count += 1
                        total.added += 1
                    else:
                        imp.updated_count += 1
                        total.updated += 1
                    db.add(AuditLog(client_id=client_id, action=action, payload=filename))
                except Exception as exc:
                    imp.error_count += 1
                    total.errors += 1
                    error_message = f"Файл «{filename}», строка {row_number}. Причина: {exc}. Что исправить: проверьте значения в этой строке и повторите загрузку."
                    total.logs.append(error_message)
                    _log_issue(db, imp.id, error_message, row_number=row_number, level="error")
                if processed == 1 or processed % 100 == 0 or processed == parsed.read_rows:
                    elapsed = monotonic() - processing_started
                    import_logger.info(
                        "Обработано строк: файл=%s, обработано=%s/%s, длительность обработки=%.3f сек",
                        filename, processed, parsed.read_rows, elapsed,
                    )
                    if progress:
                        progress("Обработка…", processed, parsed.read_rows)
            imp.skipped_count = max(parsed.read_rows - imp.added_count - imp.updated_count - imp.error_count, 0)
            total.skipped += imp.skipped_count
            stats = [
                f"Всего строк: {parsed.total_rows}",
                f"Прочитано: {parsed.read_rows}",
                f"Добавлено: {imp.added_count}",
                f"Обновлено: {imp.updated_count}",
                f"Пропущено: {imp.skipped_count}",
                f"Ошибок: {imp.error_count}",
            ]
            if parsed.read_rows and not (imp.added_count or imp.updated_count) and imp.skipped_count == parsed.read_rows:
                stats.extend(["Все записи были пропущены.", "Причина: строки прочитаны, но не были записаны в базу данных. Подробности смотрите в ошибках строк выше."])
            total.logs.extend(stats)
            for message in stats:
                _log_issue(db, imp.id, message)
        except Exception as exc:
            imp.error_count += 1
            total.errors += 1
            message = _describe_import_error(filename, exc)
            total.logs.append(message)
            _log_issue(db, imp.id, message, level="error")
        if progress:
            progress("Запись в базу…", imp.rows_count, imp.rows_count)
        commit_started = monotonic()
        db.commit()
        import_logger.info(
            "Запись в БД завершена: файл=%s, commit=%.3f сек, всего по файлу=%.3f сек",
            filename, monotonic() - commit_started, monotonic() - file_started,
        )
    import_logger.info("Импорт завершён: файлов=%s, длительность=%.3f сек", len(files), monotonic() - import_started)
    return total
